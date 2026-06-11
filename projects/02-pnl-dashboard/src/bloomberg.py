"""
Bloomberg pricing bridge via xlwings.

BDP (current prices): writes CUSIPs to Sheet1, triggers BDP formula refresh,
  reads back clean prices. Appends results to price_history.csv.

BDH (historical prices): writes one CUSIP at a time to the History sheet,
  triggers BDH formula, reads back the date/price time series.

Requirements:
  - Windows OS
  - Bloomberg Terminal running and logged in
  - xlwings + pywin32 installed
  - templates/bloomberg_prices.xlsx with Sheet1 and History sheets

Fallback: if Excel/Bloomberg is unavailable, reads from data/prices/manual_prices.csv
  or data/prices/manual_price_history.csv for historical data.
"""
import csv
import math
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd

import config
from config import get_logger

log = get_logger(__name__)

TEMPLATE_PATH = config.BLOOMBERG_TEMPLATE_PATH
PRICES_DIR = config.PRICES_DIR
MANUAL_PRICES_PATH = config.MANUAL_PRICES_PATH
MANUAL_HISTORY_PATH = config.MANUAL_HISTORY_PATH
PRICE_HISTORY_PATH = config.PRICE_HISTORY_PATH
BONDS_STATIC_PATH = config.BONDS_STATIC_PATH

_BDP_TIMEOUT = 30   # seconds
_BDH_TIMEOUT = 60   # seconds (BDH is slower)
_POLL_INTERVAL = 2

# Bloomberg ticker suffixes by asset class.
_DEFAULT_TICKER_SUFFIX = " Corp"
_KNOWN_SUFFIXES = (" Corp", " Govt", " Mtge", " Muni", " MMkt", " Index")
_BDP_SHEET = "Live MTM"   # renamed from Sheet1


def _ticker_for(cusip: str, bonds_static: dict | None = None) -> str:
    """Return full Bloomberg ticker for a CUSIP.

    Uses BondStatic.bbg_ticker if set, otherwise appends _DEFAULT_TICKER_SUFFIX.
    """
    if bonds_static:
        b = bonds_static.get(str(cusip))
        if b and b.bbg_ticker:
            return b.bbg_ticker
    return str(cusip) + _DEFAULT_TICKER_SUFFIX


def _cusip_from_ticker(ticker: str) -> str:
    """Strip Bloomberg ticker suffix to recover the bare CUSIP.

    Works for any suffix in _KNOWN_SUFFIXES; if no match the full string is
    returned as-is (handles plain CUSIPs written by older template versions).
    """
    t = str(ticker).strip()
    upper = t.upper()
    for sfx in _KNOWN_SUFFIXES:
        if upper.endswith(sfx.upper()):
            return t[: len(t) - len(sfx)].strip()
    return t


def _try_xlwings_import():
    try:
        import xlwings as xw
        return xw
    except ImportError:
        return None


# ── BDP (current prices) ─────────────────────────────────────────────────────

def _write_cusips(cusips: list[str], wb) -> None:
    """Write CUSIPs into Live MTM column A starting at row 2, forcing text format."""
    sheet = wb.sheets[_BDP_SHEET]
    last_row = sheet.range("A2").end("down").row
    if last_row >= 2:
        sheet.range(f"A2:A{max(last_row, len(cusips) + 1)}").clear_contents()
    for i, cusip in enumerate(cusips):
        cell = sheet["A" + str(i + 2)]
        cell.number_format = "@"   # force text so leading zeros are preserved
        cell.value = str(cusip)


def _is_real_price(v) -> bool:
    """True only for a finite number — excludes None, strings, and NaN (Bloomberg's #N/A response)."""
    return isinstance(v, (int, float)) and not math.isnan(v)


def _is_bloomberg_refreshing(val) -> bool:
    """Return True when a cell still holds Bloomberg's temporary 'Requesting data' placeholder.

    While Bloomberg is recalculating, cells show '#N/A Requesting data' (or variants).
    openpyxl data_only=True reads that as a plain string.  Importing at this point
    would silently write garbage; callers should block and tell the user to wait.
    """
    if not isinstance(val, str):
        return False
    lower = val.strip().lower()
    return "requesting" in lower or lower in ("#n/a", "n/a", "calculating", "loading")


def _all_prices_filled(sheet, n: int) -> bool:
    """Return True only when all price cells (col B) contain finite numeric values."""
    values = sheet.range(f"B2:B{n + 1}").value
    if values is None:
        return False
    if isinstance(values, list):
        return all(_is_real_price(v) for v in values)
    return _is_real_price(values)


def _refresh_and_read_bdp(wb, cusips: list[str]) -> dict[str, dict]:
    """Trigger BDP refresh and poll until all prices are numeric, then return results.
    Uses the original cusips list as keys to avoid Excel stripping leading zeros."""
    n_cusips = len(cusips)
    sheet = wb.sheets[_BDP_SHEET]
    wb.app.calculate()

    elapsed = 0
    while elapsed < _BDP_TIMEOUT:
        if _all_prices_filled(sheet, n_cusips):
            break
        time.sleep(_POLL_INTERVAL)
        wb.app.calculate()
        elapsed += _POLL_INTERVAL

    px_last = sheet.range(f"B2:B{n_cusips + 1}").value
    asw = sheet.range(f"C2:C{n_cusips + 1}").value
    ytm = sheet.range(f"D2:D{n_cusips + 1}").value

    if n_cusips == 1:
        px_last, asw, ytm = [px_last], [asw], [ytm]

    result = {}
    for i, cusip in enumerate(cusips):
        result[str(cusip)] = {
            "px_last": px_last[i] if _is_real_price(px_last[i]) else None,
            "asw": asw[i] if asw and _is_real_price(asw[i]) else None,
            "ytm": ytm[i] if ytm and _is_real_price(ytm[i]) else None,
        }
    return result


def prepare_template(
    cusips: list[str],
    wb_path: Path = TEMPLATE_PATH,
    bonds_static: dict | None = None,
) -> Path:
    """Write Bloomberg tickers into Live MTM col A and Static col A, then save.

    Live MTM and Static col A receive the full Bloomberg ticker (e.g.
    "912828Z78 Govt") built from BondStatic.bbg_ticker; bare CUSIPs that have
    no bbg_ticker entry default to cusip + " Corp".

    The BDP/BDH formulas in the template reference A{row} directly (no suffix
    appended in-formula), so correct pricing depends on having the right ticker
    in col A.

    After calling this the user opens the file in Excel, waits for Bloomberg to
    populate both sheets, saves and closes. Then call read_prices_from_template()
    and/or read_static_from_template().
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(wb_path))
    for sheet_name in (_BDP_SHEET, "Static"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in range(2, 52):
            ws.cell(row=row, column=1).value = None
        for i, cusip in enumerate(cusips):
            ticker = _ticker_for(cusip, bonds_static)
            ws.cell(row=i + 2, column=1).value = ticker
    wb.save(str(wb_path))
    return wb_path


def read_prices_from_template(wb_path: Path = TEMPLATE_PATH) -> dict[str, float]:
    """Read cached Bloomberg prices from a template saved after a manual Excel refresh.

    Col A contains the full Bloomberg ticker (e.g. "912828Z78 Govt"); the CUSIP
    is recovered via _cusip_from_ticker().  Col B contains PX_LAST.

    openpyxl data_only=True reads the last-calculated cell values (not formulas),
    so this only works after the user has opened the file in Excel and saved it
    with Bloomberg prices populated.
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(wb_path), data_only=True)
    sheet_name = _BDP_SHEET if _BDP_SHEET in wb.sheetnames else "Sheet1"
    ws = wb[sheet_name]

    # Scan for Bloomberg's "Requesting data" placeholder before reading anything.
    # If Bloomberg is still refreshing, block the import immediately rather than
    # importing partial/garbage data.
    refreshing_rows = [
        r for r in range(2, 52)
        if _is_bloomberg_refreshing(ws.cell(row=r, column=1).value)
        or _is_bloomberg_refreshing(ws.cell(row=r, column=2).value)
    ]
    if refreshing_rows:
        raise ValueError(
            f"Bloomberg is still refreshing Live MTM (rows {refreshing_rows[:5]}{'…' if len(refreshing_rows) > 5 else ''}). "
            "Wait until every cell shows a real value (no '#N/A Requesting data'), "
            "then Save and Close Excel before importing."
        )

    prices = {}
    for row in range(2, 52):
        raw = ws.cell(row=row, column=1).value
        price = ws.cell(row=row, column=2).value
        if not raw:
            continue
        # #N/A in the ticker cell itself → Bloomberg couldn't identify the bond
        raw_str = str(raw).strip()
        if raw_str in ("#N/A", "N/A"):
            log.debug("Live MTM row %d: ticker returned #N/A — skipped", row)
            continue
        cusip = _cusip_from_ticker(raw_str)
        # #N/A in any price field → bond has matured; Bloomberg no longer prices it
        if price is None or (isinstance(price, str) and price.strip() in ("#N/A", "N/A")):
            log.info("Live MTM: %s returned #N/A for PX_LAST — bond likely matured, skipped", cusip)
            continue
        if isinstance(price, (int, float)) and not math.isnan(float(price)):
            prices[cusip] = float(price)
    if prices:
        _save_price_snapshot(prices)
        _append_to_price_history(prices, date.today())
    return prices


def get_prices(cusips: list[str], wb_path: Path = TEMPLATE_PATH) -> dict[str, float]:
    """
    Fetch current prices via Bloomberg BDP.

    Returns {cusip: px_last}. Appends results to price_history.csv.
    Falls back to manual_prices.csv if Bloomberg is unavailable.
    """
    xw = _try_xlwings_import()

    if xw is not None:
        try:
            app = xw.App(visible=True, add_book=False)
            try:
                wb = app.books.open(str(wb_path))
                _write_cusips(cusips, wb)
                data = _refresh_and_read_bdp(wb, cusips)
                wb.save()
                wb.close()
            finally:
                app.quit()

            prices = {c: v["px_last"] for c, v in data.items() if v["px_last"] is not None}
            _save_price_snapshot(prices)
            _append_to_price_history(prices, date.today())
            return prices
        except Exception as e:
            log.warning("Bloomberg BDP failed (%s); falling back to manual prices", e)

    return _load_manual_prices(cusips)


# ── BDH (historical prices) ──────────────────────────────────────────────────

def get_historical_prices_bdh(
    cusips: list[str],
    start_date: date,
    end_date: date,
    wb_path: Path = TEMPLATE_PATH,
) -> pd.DataFrame:
    """
    Fetch historical daily prices via Bloomberg BDH for a list of CUSIPs.

    Returns a DataFrame with columns: date, cusip, px_last.
    Appends results to price_history.csv.
    Falls back to manual_price_history.csv if Bloomberg is unavailable.

    Uses the "History" sheet in bloomberg_prices.xlsx:
        A1: ticker string (e.g. "037833100 Corp")
        B1: start date  ("YYYYMMDD")
        C1: end date    ("YYYYMMDD")
        A3: =BDH(A1,"PX_LAST",B1,C1,"Fill","0","Dates","0")
              ↑ BDH fills downward: col A = dates, col B = prices
    """
    xw = _try_xlwings_import()

    if xw is not None:
        try:
            records = _fetch_bdh_via_excel(cusips, start_date, end_date, wb_path, xw)
            if records:
                df = pd.DataFrame(records)
                _append_df_to_price_history(df)
                return df
        except Exception as e:
            log.warning("Bloomberg BDH failed (%s); falling back to manual history", e)

    return _load_manual_price_history(cusips, start_date, end_date)


def _fetch_bdh_via_excel(cusips, start_date, end_date, wb_path, xw,
                         bonds_static=None) -> list[dict]:
    # Dates as mm/dd/yyyy — consistent with the History sheet template layout.
    start_str = start_date.strftime("%m/%d/%Y")
    end_str   = end_date.strftime("%m/%d/%Y")

    app = xw.App(visible=False, add_book=False)
    records = []
    try:
        wb = app.books.open(str(wb_path))
        sheet = wb.sheets["History"]

        for cusip in cusips:
            sheet.clear_contents()

            # Write parameters for single-CUSIP live call (A1/B1/C1 layout).
            sheet["A1"].value = _ticker_for(cusip, bonds_static)
            sheet["B1"].value = start_str
            sheet["C1"].value = end_str

            # @BDH formula: dates fill col A from row 3, prices fill col B.
            sheet["A3"].value = '=@BDH(A1,"PX_LAST",B1,C1,"fill=p","days=a")'

            # Trigger and wait for Bloomberg to populate.
            app.calculate()
            elapsed = 0
            while elapsed < _BDH_TIMEOUT:
                val = sheet["B3"].value
                if val is not None and val != "":
                    break
                time.sleep(_POLL_INTERVAL)
                app.calculate()
                elapsed += _POLL_INTERVAL

            # Read returned time series.
            data_range = sheet.range("A3").expand().value
            if data_range is None:
                continue
            # Single-row result comes back as a flat list, not a list of lists.
            if not isinstance(data_range[0], list):
                data_range = [data_range]

            for row in data_range:
                if row and row[0] and row[1]:
                    row_date = row[0].date() if hasattr(row[0], "date") else row[0]
                    records.append({
                        "date": row_date,
                        "cusip": cusip,
                        "px_last": float(row[1]),
                    })

        wb.close()
    finally:
        app.quit()

    return records


# ── price_history.csv helpers ─────────────────────────────────────────────────

def _append_to_price_history(prices: dict[str, float], as_of: date) -> None:
    """Append a dict of {cusip: px_last} for a single date to price_history.csv."""
    rows = [{"date": as_of.isoformat(), "cusip": c, "px_last": px} for c, px in prices.items()]
    _append_df_to_price_history(pd.DataFrame(rows))


def _append_df_to_price_history(df: pd.DataFrame) -> None:
    """Append a DataFrame (date, cusip, px_last) to price_history.csv, deduplicating."""
    path = PRICE_HISTORY_PATH

    if path.exists() and path.stat().st_size > 0:
        existing = pd.read_csv(path, dtype={"cusip": str})
        combined = pd.concat([existing, df], ignore_index=True)
        combined["date"] = pd.to_datetime(combined["date"]).dt.date.astype(str)
        combined = combined.drop_duplicates(subset=["date", "cusip"], keep="last")
        combined.sort_values(["date", "cusip"]).to_csv(path, index=False)
    else:
        df["date"] = pd.to_datetime(df["date"]).dt.date.astype(str)
        df.sort_values(["date", "cusip"]).to_csv(path, index=False)


def load_price_history(path: Path = PRICE_HISTORY_PATH) -> pd.DataFrame:
    """Load full price history; returns empty DataFrame if file doesn't exist."""
    if not Path(path).exists() or Path(path).stat().st_size == 0:
        return pd.DataFrame(columns=["date", "cusip", "px_last"])
    df = pd.read_csv(path, dtype={"cusip": str})
    df["date"] = pd.to_datetime(df["date"])
    df["px_last"] = pd.to_numeric(df["px_last"], errors="coerce")
    return df


# ── snapshot helpers (daily timestamped files) ────────────────────────────────

def _save_price_snapshot(prices: dict[str, float]) -> None:
    Path(PRICES_DIR).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = PRICES_DIR / f"prices_{ts}.csv"
    with open(out_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["cusip", "px_last", "date"])
        today = date.today().isoformat()
        for cusip, px in prices.items():
            writer.writerow([str(cusip), px, today])


def load_latest_prices() -> dict[str, float]:
    """Load the most recently saved BDP snapshot (Bloomberg or manual)."""
    snapshots = sorted(PRICES_DIR.glob("prices_*.csv"))
    if snapshots:
        prices = {}
        with open(snapshots[-1], newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                prices[row["cusip"]] = float(row["px_last"])
        return prices
    return _load_manual_prices(cusips=[])


# ── history template (manual two-step BDH workflow) ──────────────────────────

_BACKFILL_STATE_PATH = PRICES_DIR / "backfill_state.json"
_BDH_BLOCK_ROWS = 510   # rows reserved per CUSIP block (ample for ~500 business days)


def _load_backfill_state() -> dict:
    if _BACKFILL_STATE_PATH.exists():
        import json
        return json.loads(_BACKFILL_STATE_PATH.read_text())
    return {}


def _save_backfill_state(state: dict) -> None:
    import json
    _BACKFILL_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    _BACKFILL_STATE_PATH.write_text(json.dumps(state, indent=2, default=str))


def last_priced_date() -> date | None:
    """Return the latest date in price_history.csv, or None if the file is empty."""
    if not PRICE_HISTORY_PATH.exists():
        return None
    try:
        df = pd.read_csv(PRICE_HISTORY_PATH, dtype={"cusip": str})
        if df.empty:
            return None
        return pd.to_datetime(df["date"]).max().date()
    except Exception:
        return None


def _prev_business_day(d: date) -> date:
    """Most recent business day strictly before d (skips Sat/Sun)."""
    d = d - timedelta(days=1)
    while d.weekday() >= 5:  # 5=Sat, 6=Sun
        d -= timedelta(days=1)
    return d


def find_price_gaps(
    all_trades: pd.DataFrame,
    end_date: date | None = None,
) -> list[tuple[str, date, date]]:
    """Return (cusip, gap_start, gap_end) ranges needed for MTM but absent from price_history.csv.

    Algorithm:
      1. Walk each CUSIP's signed-nominal trade stream to find every hold interval
         (date range where net_nominal != 0).  A CUSIP re-bought after a full close
         produces two separate intervals.
      2. Compare every business day in each interval against price_history.csv.
      3. Collapse each run of missing business days into a single (start, end) range.

    Callers pass this directly to prepare_history_template() — no manual date
    picking needed even after backdated trade edits or new CUSIPs are added.
    """
    if all_trades.empty:
        return []

    if end_date is None:
        # BDH returns end-of-day closes — today's close does not exist yet
        # (today's price comes from the Live MTM intraday BDP sheet). Cap at the
        # previous business day so `today` never enters gap detection, which
        # would otherwise create an invalid BDH(start=today, end=yesterday).
        end_date = _prev_business_day(date.today())

    # ── step 1: compute hold intervals per CUSIP ─────────────────────────────
    hold_intervals: dict[str, list[tuple[date, date]]] = {}

    for cusip, group in all_trades.groupby("cusip"):
        group = group.sort_values("trade_date").reset_index(drop=True)
        cusip = str(cusip)
        running = 0.0
        interval_start: date | None = None
        intervals: list[tuple[date, date]] = []

        for _, row in group.iterrows():
            was_flat = abs(running) < 0.01
            running += float(row["nominal"])   # signed: +buy, –sell
            is_flat = abs(running) < 0.01
            trade_dt: date = (
                row["trade_date"].date()
                if hasattr(row["trade_date"], "date")
                else row["trade_date"]
            )

            if was_flat and not is_flat:
                interval_start = trade_dt
            elif not was_flat and is_flat and interval_start is not None:
                intervals.append((interval_start, trade_dt))
                interval_start = None

        if interval_start is not None:
            intervals.append((interval_start, end_date))

        if intervals:
            hold_intervals[cusip] = intervals

    # ── step 2: load what we already have ────────────────────────────────────
    have: set[tuple[str, str]] = set()
    if PRICE_HISTORY_PATH.exists() and PRICE_HISTORY_PATH.stat().st_size > 0:
        try:
            hist = pd.read_csv(PRICE_HISTORY_PATH, dtype={"cusip": str})
            have = set(zip(hist["cusip"].astype(str), hist["date"].astype(str)))
        except Exception:
            pass

    # ── step 3: find gaps (missing business days) per interval ───────────────
    result: list[tuple[str, date, date]] = []

    for cusip, intervals in hold_intervals.items():
        for (ivl_start, ivl_end) in intervals:
            bdays = [
                d.date()
                for d in pd.bdate_range(start=ivl_start, end=min(ivl_end, end_date))
            ]
            missing = [d for d in bdays if (cusip, d.isoformat()) not in have]
            if not missing:
                continue

            # Collapse consecutive missing business days into contiguous ranges.
            # Two adjacent elements of `missing` are consecutive business days
            # iff there is no business day between them — equivalently, they
            # were neighbours in the original `bdays` list.
            bday_set = set(bdays)
            gap_start = missing[0]
            prev = missing[0]
            for d in missing[1:]:
                # Check whether prev and d are adjacent in the full bday sequence.
                # They are NOT adjacent if there is at least one bday between them
                # (i.e. a date we already have in price_history).
                nxt = prev + timedelta(days=1)
                while nxt not in bday_set and nxt < d:
                    nxt += timedelta(days=1)
                if nxt != d:
                    # Gap interrupted — emit the completed range and start a new one
                    result.append((cusip, gap_start, prev))
                    gap_start = d
                prev = d
            result.append((cusip, gap_start, prev))

    return result


def prepare_history_template(
    cusip_ranges: list[tuple[str, date, date]],
    wb_path: Path = TEMPLATE_PATH,
    bonds_static: dict | None = None,
) -> Path:
    """Write the History sheet with one column pair per CUSIP using @BDH formulas.

    Layout:
      Row 1:  odd cols  = tickers  (A1, C1, E1, …)
              bond-0 even col = start date  (B1)
      Row 2:  A2 = shared end date (today−1, mm/dd/yyyy)
              B2 = "PX_LAST"  (shared field label)
              bond-1+ even cols = start dates  (D2, F2, …)
      Row 3:  odd cols = @BDH formulas — Bloomberg fills dates (odd) and
              prices (even) downward from row 4

    Multiple gap ranges for the same CUSIP are collapsed: the earliest start
    date is used (find_price_gaps already returns only missing ranges, so
    min(start) = the earliest date not yet in price_history.csv).

    After this call the user opens the file in Excel, waits for Bloomberg to
    populate every formula, saves and closes. Then call read_history_from_template().
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    end_date = _prev_business_day(date.today())
    end_str  = end_date.strftime("%m/%d/%Y")

    # Collapse multiple ranges per CUSIP: keep earliest start date.
    # Skip CUSIPs whose earliest gap starts AFTER end_date — BDH needs
    # start <= end to produce data; these bonds are priced via Live MTM only.
    cusip_starts: dict[str, date] = {}
    for cusip, start, _end in cusip_ranges:
        cusip = str(cusip)
        if start > end_date:
            log.info(
                "prepare_history_template: %s gap starts %s > end %s — skipped (use Live MTM)",
                cusip, start, end_date,
            )
            continue
        if cusip not in cusip_starts or start < cusip_starts[cusip]:
            cusip_starts[cusip] = start

    wb = load_workbook(str(wb_path))
    ws = wb["History"]
    ws.delete_rows(1, max(ws.max_row, 1))

    # Shared cells — written once, referenced by all BDH formulas.
    ws["A2"] = end_str
    ws["B2"] = "PX_LAST"

    ordered_cusips = list(cusip_starts.keys())
    for i, cusip in enumerate(ordered_cusips):
        col_data   = 2 * i + 1                   # A=1, C=3, E=5, …
        col_param  = 2 * i + 2                   # B=2, D=4, F=6, …
        col_letter = get_column_letter(col_data)
        start_str  = cusip_starts[cusip].strftime("%m/%d/%Y")
        ticker     = _ticker_for(cusip, bonds_static)

        # Ticker — row 1, odd column.
        cell = ws.cell(row=1, column=col_data, value=ticker)
        cell.font = Font(bold=True)

        # Start date: B1 for bond 0; D2, F2, … for subsequent bonds.
        if i == 0:
            ws.cell(row=1, column=col_param, value=start_str)   # B1
            start_ref = "B1"
        else:
            ws.cell(row=2, column=col_param, value=start_str)   # D2, F2, …
            start_ref = f"{get_column_letter(col_param)}2"

        # @BDH formula — Bloomberg fills dates into col_data and prices into
        # col_param starting at row 4 (row 3 is the formula cell itself).
        formula = f'=@BDH({col_letter}1,B2,{start_ref},A2,"fill=p","days=a")'
        ws.cell(row=3, column=col_data, value=formula)

    wb.save(str(wb_path))

    prev = _load_backfill_state()
    _save_backfill_state({
        "ordered_cusips": ordered_cusips,
        "end_date": end_str,
        "prepared_at": datetime.now().isoformat(),
        "last_date": prev.get("last_date"),
        "last_run": prev.get("last_run"),
    })

    return wb_path


def read_history_from_template(wb_path: Path = TEMPLATE_PATH) -> pd.DataFrame:
    """Read column-pair BDH data from the History sheet after Bloomberg has populated it.

    Each CUSIP occupies two columns (dates in odd, prices in even), data from row 4
    downward (row 3 is the @BDH formula cell).  The ordered CUSIP list stored in
    backfill_state.json maps column positions back to CUSIPs.

    Returns a DataFrame (date, cusip, px_last). Appends to price_history.csv
    and updates backfill_state.json with the latest date successfully imported.
    """
    from openpyxl import load_workbook

    state = _load_backfill_state()
    ordered_cusips: list[str] = state.get("ordered_cusips", [])
    if not ordered_cusips:
        return pd.DataFrame(columns=["date", "cusip", "px_last"])

    wb = load_workbook(str(wb_path), data_only=True)
    ws = wb["History"]
    max_row = ws.max_row or 3

    # Check for Bloomberg's refreshing placeholder in the first data rows.
    n_data_cols = len(ordered_cusips) * 2
    refreshing_cells = [
        f"row {r} col {c}"
        for r in range(3, min(max_row + 1, 8))   # sample a few rows; full scan is slow
        for c in range(1, n_data_cols + 1)
        if _is_bloomberg_refreshing(ws.cell(row=r, column=c).value)
    ]
    if refreshing_cells:
        raise ValueError(
            f"Bloomberg is still refreshing History sheet ({len(refreshing_cells)} cell(s) sampled, "
            f"e.g. {refreshing_cells[0]}). "
            "Wait until BDH blocks show real date/price rows (no '#N/A Requesting data'), "
            "then Save and Close Excel before importing."
        )

    records = []
    for i, cusip in enumerate(ordered_cusips):
        col_data  = 2 * i + 1   # odd column: dates
        col_price = 2 * i + 2   # even column: prices

        for r in range(4, max_row + 1):   # row 3 is the formula; data starts at 4
            date_val  = ws.cell(row=r, column=col_data).value
            price_val = ws.cell(row=r, column=col_price).value
            if not date_val:
                break
            if hasattr(date_val, "date"):
                d = date_val.date()
            else:
                # Fallback: Bloomberg may return mm/dd/yyyy string
                try:
                    d = datetime.strptime(str(date_val), "%m/%d/%Y").date()
                except ValueError:
                    try:
                        d = datetime.strptime(str(int(float(str(date_val)))), "%Y%m%d").date()
                    except (ValueError, TypeError):
                        break
            if isinstance(price_val, (int, float)) and not math.isnan(float(price_val)):
                records.append({
                    "date": d.isoformat(),
                    "cusip": cusip,
                    "px_last": float(price_val),
                })

    if not records:
        return pd.DataFrame(columns=["date", "cusip", "px_last"])

    df = pd.DataFrame(records)
    _append_df_to_price_history(df)

    state["last_date"] = df["date"].max()
    state["last_run"] = datetime.now().isoformat()
    _save_backfill_state(state)

    return df


# ── bond static template (BDP reference data) ────────────────────────────────

# Map Bloomberg DAY_CNT_DES strings to our canonical day-count set.
_DAYCOUNT_MAP: dict[str, str] = {
    "ACT/360": "Act/360",
    "ACTUAL/360": "Act/360",
    "A/360": "Act/360",
    "ACT/365": "Act/365",
    "ACTUAL/365": "Act/365",
    "A/365": "Act/365",
    "ACT/ACT": "Act/Act",
    "ACTUAL/ACTUAL": "Act/Act",
    "A/A": "Act/Act",
    "ACT/ACT (ISMA)": "Act/Act",
    "ACT/ACT (ICMA)": "Act/Act",
    "ACTUAL/ACTUAL (ISMA)": "Act/Act",
    "30/360": "30/360",
    "30/360 BOND": "30/360",
    "30/360 ISDA": "30/360",
    "ISMA-30/360": "30/360",
    "ISMA-30/360 NONEOM": "30/360",
    "ISMA 30/360": "30/360",
    "30E/360": "30/360",
    "BOND BASIS": "30/360",
}


def _normalize_day_count(raw) -> str | None:
    if raw is None:
        return None
    upper = str(raw).strip().upper()
    result = _DAYCOUNT_MAP.get(upper)
    if result is None:
        log.warning("Unknown Bloomberg day count %r — leaving blank", raw)
    return result


def read_static_from_template(wb_path: Path = TEMPLATE_PATH) -> pd.DataFrame:
    """Read bond reference data from the Static sheet of a Bloomberg-populated template.

    Column order matches bonds_static.csv exactly:
      cusip, name, currency, country, coupon_rate, coupon_frequency,
      day_count_convention, maturity_date, first_coupon_date

    The formula in the template already divides CPN by 100, so coupon_rate
    arrives as a decimal (e.g. 0.055 for 5.5%).
    Bloomberg dates come back as Python datetime objects; stored as ISO strings.
    Day-count strings are normalised to {Act/360, Act/365, 30/360}.
    """
    from openpyxl import load_workbook
    from data_io import BONDS_COLUMNS

    wb = load_workbook(str(wb_path), data_only=True)
    if "Static" not in wb.sheetnames:
        return pd.DataFrame(columns=BONDS_COLUMNS)

    ws = wb["Static"]
    # Read headers dynamically — stop at first empty cell so the reader
    # automatically handles templates with more or fewer columns than expected.
    headers: list[str | None] = []
    for c in range(1, ws.max_column + 2):
        h = ws.cell(row=1, column=c).value
        if h is None:
            break
        headers.append(str(h))

    if not headers:
        log.warning("Static sheet has no headers — was the template regenerated?")
        return pd.DataFrame(columns=BONDS_COLUMNS)

    # Scan every data cell for Bloomberg's refreshing placeholder before reading.
    n_cols = len(headers)
    refreshing_cells = [
        f"row {r} col {c}"
        for r in range(2, 52)
        for c in range(1, n_cols + 1)
        if _is_bloomberg_refreshing(ws.cell(row=r, column=c).value)
    ]
    if refreshing_cells:
        raise ValueError(
            f"Bloomberg is still refreshing Static sheet ({len(refreshing_cells)} cell(s), "
            f"e.g. {refreshing_cells[0]}). "
            "Wait until every cell shows a real value (no '#N/A Requesting data'), "
            "then Save and Close Excel before importing."
        )

    records = []
    for row in range(2, 52):
        raw_a = ws.cell(row=row, column=1).value
        if not raw_a:
            continue

        # Col A may hold a full ticker ("912828Z78 Govt") or a bare CUSIP from
        # older templates — _cusip_from_ticker handles both.
        rec: dict = {"cusip": _cusip_from_ticker(str(raw_a))}
        for col_idx, field in enumerate(headers[1:], 2):
            if not field:
                continue
            val = ws.cell(row=row, column=col_idx).value

            if field in ("name", "currency", "country"):
                # Plain string fields: None (openpyxl renders #N/A errors as None)
                # or non-string values → empty string.
                val = "" if val is None else str(val).strip()
                if val.upper() in ("#N/A", "N/A"):
                    val = ""
            elif field == "day_count_convention":
                val = _normalize_day_count(val)
            elif field == "maturity_date":
                if hasattr(val, "date"):
                    val = val.date().isoformat()
                elif val is not None and str(val).strip() not in ("#N/A", "N/A", ""):
                    val = str(val)
                else:
                    val = None  # Bloomberg #N/A → can't import this bond
            elif field == "first_coupon_date":
                # Bloomberg returns #N/A (None from xlwings, or a string) when
                # a bond has no distinct first coupon (zero-coupon / bullet).
                # Sentinel None here; merge_bonds_static will fill it with maturity.
                if hasattr(val, "date"):
                    val = val.date().isoformat()
                elif val is not None and str(val).strip() not in ("#N/A", "N/A", ""):
                    val = str(val)
                else:
                    val = None  # will be filled with maturity_date in merge step
            elif field == "coupon_rate":
                val = float(val) if val is not None else None
            elif field == "coupon_frequency":
                # Bloomberg returns #N/A (None) for non-coupon-bearing bonds → 0
                if val is None or (isinstance(val, str) and val.strip() in ("#N/A", "N/A", "")):
                    val = 0
                else:
                    try:
                        val = int(round(float(val)))
                    except (ValueError, TypeError):
                        val = 0

            rec[field] = val

        # maturity_date is required — skip any row where Bloomberg returned #N/A
        if rec.get("maturity_date") is None:
            log.warning(
                "Static row for %s: maturity_date is missing/N/A — "
                "bond may have matured or the ticker is incorrect. Row skipped.",
                rec.get("cusip", "?"),
            )
            continue
        records.append(rec)

    if not records:
        return pd.DataFrame(columns=BONDS_COLUMNS)

    return pd.DataFrame(records).reindex(columns=BONDS_COLUMNS)


def merge_bonds_static(fetched_df: pd.DataFrame) -> tuple[int, int]:
    """Fill gaps in bonds_static.csv using Bloomberg-fetched reference data.

    Existing values are preserved — only NaN / blank cells are filled.
    CUSIPs not yet in bonds_static.csv are added entirely.

    Returns (n_new_cusips, n_fields_filled).
    Raises ValueError (propagated from save_bonds_static) if any merged row
    still fails BondStatic validation after the merge.
    """
    from data_io import save_bonds_static, BONDS_COLUMNS

    if fetched_df.empty:
        return (0, 0)

    if BONDS_STATIC_PATH.exists() and BONDS_STATIC_PATH.stat().st_size > 0:
        existing = pd.read_csv(BONDS_STATIC_PATH, dtype={"cusip": str})
    else:
        existing = pd.DataFrame(columns=BONDS_COLUMNS)

    existing["cusip"] = existing["cusip"].astype(str)
    fetched = fetched_df.reindex(columns=BONDS_COLUMNS).copy()
    fetched["cusip"] = fetched["cusip"].astype(str)

    new_cusips = set(fetched["cusip"]) - set(existing["cusip"])

    # combine_first: existing values win; gaps filled from fetched
    merged = (
        existing.set_index("cusip")
        .combine_first(fetched.set_index("cusip"))
        .reset_index()
    )

    # first_coupon_date sentinel (None) from Bloomberg #N/A → use maturity_date
    if "first_coupon_date" in merged.columns and "maturity_date" in merged.columns:
        mask = merged["first_coupon_date"].isna()
        merged.loc[mask, "first_coupon_date"] = merged.loc[mask, "maturity_date"]

    # Count cells that were NaN before but are now filled
    pre = existing.set_index("cusip").reindex(merged.set_index("cusip").index)
    post = merged.set_index("cusip")
    fields_filled = int((post.notna() & pre.isna()).sum().sum())

    save_bonds_static(merged)
    return (len(new_cusips), fields_filled)


def open_in_excel(path: Path) -> bool:
    """Best-effort: open a file in Excel/the system's default spreadsheet app.

    Works on Windows (os.startfile), macOS (open), and Linux (xdg-open).
    Returns True if the launch succeeded, False if it failed or the platform
    couldn't be detected — caller should fall back to showing the file path.
    """
    import sys
    import subprocess
    try:
        if sys.platform == "win32":
            import os
            os.startfile(str(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
        return True
    except Exception as exc:
        log.warning("Could not open %s in Excel: %s", path, exc)
        return False


# ── manual fallbacks ──────────────────────────────────────────────────────────

def _load_manual_prices(cusips: list[str]) -> dict[str, float]:
    """Most recent price per CUSIP from manual_prices.csv."""
    if not MANUAL_PRICES_PATH.exists():
        return {}

    prices: dict[str, tuple[str, float]] = {}
    with open(MANUAL_PRICES_PATH, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cusip = row["cusip"]
            if not cusips or cusip in cusips:
                existing_date = prices.get(cusip, ("", 0.0))[0]
                if row["date"] >= existing_date:
                    prices[cusip] = (row["date"], float(row["px_last"]))

    return {cusip: v[1] for cusip, v in prices.items()}


def _load_manual_price_history(cusips, start_date, end_date) -> pd.DataFrame:
    if not MANUAL_HISTORY_PATH.exists():
        return pd.DataFrame(columns=["date", "cusip", "px_last"])

    df = pd.read_csv(MANUAL_HISTORY_PATH, dtype={"cusip": str})
    df["date"] = pd.to_datetime(df["date"])
    mask = (df["date"] >= pd.Timestamp(start_date)) & (df["date"] <= pd.Timestamp(end_date))
    if cusips:
        mask &= df["cusip"].isin(cusips)
    return df[mask].reset_index(drop=True)
