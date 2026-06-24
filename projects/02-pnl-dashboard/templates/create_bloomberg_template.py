"""
Run once to create bloomberg_prices.xlsx with Live MTM, Static, and History sheets.

Usage:  python templates/create_bloomberg_template.py

IMPORTANT — regenerate the template any time this script changes.
Col A in Live MTM and Static holds the FULL Bloomberg ticker written by the
dashboard (e.g. "912828Z78 Govt").  BDP formulas reference A{row} directly —
no ticker suffix is appended in the formula itself.
"""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Single source of truth for the row count lives in src/config.py so the
# generator and the importer can never drift out of sync.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import config  # noqa: E402

OUTPUT = Path(__file__).parent / "bloomberg_prices.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
N_ROWS = config.BLOOMBERG_TEMPLATE_ROWS   # securities on Live MTM / Static sheets


def _header(ws, col, text):
    cell = ws.cell(row=1, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")


def create_live_mtm_sheet(wb: Workbook) -> None:
    """Live MTM sheet — current prices via BDP.

    Col A: full Bloomberg ticker written by the dashboard (e.g. "912828Z78 Govt").
    Formulas reference A{row} directly — no suffix appended in-formula — so
    prices work for any asset class (Corp, Govt, Mtge, Muni, etc.).
    """
    ws = wb.active
    ws.title = "Live MTM"

    for col, header in enumerate(["BBG Ticker", "Last Price (PX_LAST)", "ASW Spread", "YTM (Mid)"], 1):
        _header(ws, col, header)

    fields = ["PX_LAST", "YAS_ASW_SPREAD", "YLD_YTM_MID"]
    for row in range(2, N_ROWS + 2):
        for col_offset, field in enumerate(fields, 1):
            ws.cell(row=row, column=col_offset + 1,
                    value=f'=BDP(A{row},"{field}")')

    ws.column_dimensions["A"].width = 22
    for col in range(2, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22


def create_static_sheet(wb: Workbook) -> None:
    """Static sheet — bond reference data via BDP.

    Col A: full Bloomberg ticker (same as Live MTM col A) written by dashboard.
    All other columns are BDP formulas referencing A{row} directly.

    Field mapping notes:
      - coupon_rate: Bloomberg CPN is a percent (e.g. 5.5); divided by 100 in
        the formula so the cell holds 0.055 (the decimal bonds_static wants).
      - day_count_convention: DAY_CNT_DES returns Bloomberg's raw string;
        the importer normalises it to {Act/360, Act/365, 30/360}.
      - maturity_date / first_coupon_date: returned as Excel date values.
    """
    ws = wb.create_sheet("Static")

    # (header, BDP field expression keyed off A{row}; None = ticker col written by caller)
    # `market` (Local/Global) has no single clean BDP mnemonic — it is derived
    # from currency-vs-country-of-risk in classification.py, so it is left as a
    # manual/derived column (None expr).
    columns = [
        ("bbg_ticker",          None),
        ("name",                'BDP(A{r},"SECURITY_DES")'),
        ("currency",            'BDP(A{r},"CRNCY")'),
        ("country",             'BDP(A{r},"CNTRY_OF_RISK")'),
        ("coupon_rate",         'BDP(A{r},"CPN")/100'),
        ("coupon_frequency",    'BDP(A{r},"CPN_FREQ")'),
        ("day_count_convention",'BDP(A{r},"DAY_CNT_DES")'),
        ("maturity_date",       'BDP(A{r},"MATURITY")'),
        ("first_coupon_date",   'BDP(A{r},"FIRST_CPN_DT")'),
        # ── enterprise classification fields ──────────────────────────────────
        ("instrument_type",     'BDP(A{r},"MARKET_SECTOR_DES")'),
        ("issuer",              'BDP(A{r},"ISSUER")'),
        ("country_of_risk",     'BDP(A{r},"CNTRY_OF_RISK")'),
        ("sector",              'BDP(A{r},"INDUSTRY_SECTOR")'),
        ("seniority",           'BDP(A{r},"PAYMENT_RANK")'),
        ("market",              None),
        ("rating_sp",           'BDP(A{r},"RTG_SP")'),
        ("rating_moody",        'BDP(A{r},"RTG_MOODY")'),
        ("rating_fitch",        'BDP(A{r},"RTG_FITCH")'),
    ]

    for col, (header, _) in enumerate(columns, 1):
        _header(ws, col, header)

    for row in range(2, N_ROWS + 2):
        for col, (_, expr) in enumerate(columns, 1):
            if expr is None:
                continue   # column A holds the ticker, written by the caller
            ws.cell(row=row, column=col,
                    value="=" + expr.format(r=row))

    ws.column_dimensions["A"].width = 22
    for col in range(2, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def create_history_sheet(wb: Workbook) -> None:
    """History sheet — time-series prices via BDH (column-pair layout).

    Populated at runtime by prepare_history_template() in bloomberg.py.
    Layout (one column pair per bond):

      Row 1:  A1=Ticker1  B1=StartDate1  C1=Ticker2  (empty)  E1=Ticker3  ...
      Row 2:  A2=EndDate  B2=PX_LAST     (empty)     D2=Start2 (empty)   F2=Start3 ...
      Row 3:  =@BDH(...)  [prices↓]      =@BDH(...)  [prices↓] ...
      Row 4+: date/price data filled by Bloomberg

    Dates are mm/dd/yyyy.  A2 (end date = today−1) is shared by all formulas.
    B2 ("PX_LAST") is the shared field label referenced by every BDH formula.
    """
    ws = wb.create_sheet("History")

    ws["A1"] = "BBG Ticker"
    ws["A2"] = "End Date (mm/dd/yyyy)"
    ws["B2"] = "PX_LAST"
    for cell in (ws["A1"], ws["A2"], ws["B2"]):
        cell.font = Font(bold=True)

    for i in range(N_ROWS):   # placeholder widths, one column-pair per bond
        ws.column_dimensions[get_column_letter(2 * i + 1)].width = 22
        ws.column_dimensions[get_column_letter(2 * i + 2)].width = 14


def main():
    wb = Workbook()
    create_live_mtm_sheet(wb)
    create_static_sheet(wb)
    create_history_sheet(wb)
    wb.save(OUTPUT)
    print(f"Template saved → {OUTPUT}")
    print("Live MTM: BDP current prices — col A = full BBG ticker (e.g. '912828Z78 Govt')")
    print("Static:   BDP bond reference data — col A = full BBG ticker")
    print("History:  BDH time-series blocks written at runtime by the dashboard")
    print()
    print("NOTE: After saving, use the dashboard '① Prepare & open template' button")
    print("to write the correct tickers into col A before Bloomberg populates.")


if __name__ == "__main__":
    main()
