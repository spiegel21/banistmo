"""Tests for the Bloomberg template row capacity (no 50-row cap)."""
import pytest
from openpyxl import load_workbook

import bloomberg
import config


def test_template_row_constant_is_at_least_100():
    assert config.BLOOMBERG_TEMPLATE_ROWS >= 100


def test_committed_template_sized_to_constant():
    """Live MTM and Static carry BDP formulas for every configured row."""
    wb = load_workbook(config.BLOOMBERG_TEMPLATE_PATH)
    for sheet in ("Live MTM", "Static"):
        ws = wb[sheet]
        formula_rows = [
            c.row
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ]
        # rows 2 .. ROWS+1 inclusive
        assert max(formula_rows) == config.BLOOMBERG_TEMPLATE_ROWS + 1
        assert min(formula_rows) == 2


def test_prepare_template_rejects_more_securities_than_rows():
    """Bonds beyond the template size must fail loudly, never be dropped silently."""
    too_many = [f"{i:09d}" for i in range(config.BLOOMBERG_TEMPLATE_ROWS + 1)]
    with pytest.raises(ValueError, match="exceed the template"):
        bloomberg.prepare_template(too_many)
