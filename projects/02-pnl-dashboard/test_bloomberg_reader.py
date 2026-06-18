"""
Bloomberg Template Reader Diagnostic
=====================================
Shows exactly what openpyxl data_only=True reads from each sheet in
bloomberg_prices.xlsx, cross-referenced against price_history.csv to
explain which bonds appear in the History BDH sheet and why.

Run from the project root:
    python test_bloomberg_reader.py
"""
import json
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, "src")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import config

TEMPLATE       = Path("templates/bloomberg_prices.xlsx")
PRICE_HISTORY  = Path("data/price_history.csv")
BACKFILL_STATE = Path("data/prices/backfill_state.json")

print("Template exists     :", TEMPLATE.exists())
print("Price history exists:", PRICE_HISTORY.exists())
print()

# ── 1. Live MTM sheet ────────────────────────────────────────────────────────
wb = load_workbook(str(TEMPLATE), data_only=True)
print("Sheets:", wb.sheetnames)

ws_live = wb["Live MTM"]
print(f"\n=== Live MTM  (max_row={ws_live.max_row}) ===")
print(f"{'Row':<5} {'Ticker/CUSIP':<25} {'PX_LAST':<18} Notes")
print("-" * 75)
for r in range(1, ws_live.max_row + 1):
    c1 = ws_live.cell(r, 1).value
    c2 = ws_live.cell(r, 2).value
    if c1 is None and c2 is None:
        break
    notes = ""
    if isinstance(c2, str) and ("n/a" in c2.lower() or "request" in c2.lower()):
        notes = "⚠ Bloomberg still refreshing"
    elif c2 is None and r > 1:
        notes = "← None (matured or formula not cached)"
    print(f"{r:<5} {str(c1):<25} {str(c2):<18} {notes}")

# ── 2. Static sheet ──────────────────────────────────────────────────────────
ws_static = wb["Static"]
print(f"\n=== Static  (max_row={ws_static.max_row}) ===")
headers: list[str] = []
for c in range(1, ws_static.max_column + 2):
    h = ws_static.cell(1, c).value
    if h is None:
        break
    headers.append(str(h))
print("Headers:", headers)

rows_static = []
for r in range(2, ws_static.max_row + 1):
    row = {h: ws_static.cell(r, ci + 1).value for ci, h in enumerate(headers)}
    if all(v is None for v in row.values()):
        break
    rows_static.append(row)
if rows_static:
    print(pd.DataFrame(rows_static).to_string(index=False))
else:
    print("(empty — no data rows)")

# ── 3. History sheet ─────────────────────────────────────────────────────────
ws_hist = wb["History"]
print(f"\n=== History  (max_row={ws_hist.max_row}, max_col={ws_hist.max_column}) ===")
print("Row 1 (tickers):")
for c in range(1, ws_hist.max_column + 1):
    v = ws_hist.cell(1, c).value
    if v is not None:
        print(f"  col {c} ({get_column_letter(c)}): {v}")
print("Row 2 (end date / start dates / field):")
for c in range(1, ws_hist.max_column + 1):
    v = ws_hist.cell(2, c).value
    if v is not None:
        print(f"  col {c} ({get_column_letter(c)}): {v}")
print("Row 3 (BDH formula / first data row):")
for c in range(1, ws_hist.max_column + 1):
    v = ws_hist.cell(3, c).value
    if v is not None:
        print(f"  col {c} ({get_column_letter(c)}): {v}")
print("Sample data rows 4–8:")
for r in range(4, 9):
    vals = [
        f"col{c}={ws_hist.cell(r, c).value}"
        for c in range(1, ws_hist.max_column + 1)
        if ws_hist.cell(r, c).value is not None
    ]
    if vals:
        print(f"  row {r}: {' | '.join(vals)}")

# ── 4. Backfill state ────────────────────────────────────────────────────────
print("\n=== Backfill state (data/prices/backfill_state.json) ===")
if BACKFILL_STATE.exists():
    state = json.loads(BACKFILL_STATE.read_text())
    print(f"  ordered_cusips : {state.get('ordered_cusips')}")
    print(f"  end_date       : {state.get('end_date')}")
    print(f"  prepared_at    : {state.get('prepared_at')}")
    print(f"  last_date      : {state.get('last_date')}")
else:
    print("  NOT FOUND — History sheet has never been prepared/imported.")

# ── 5. Price history coverage ────────────────────────────────────────────────
print("\n=== price_history.csv coverage per CUSIP ===")
if PRICE_HISTORY.exists():
    ph = pd.read_csv(PRICE_HISTORY, dtype={"cusip": str})
    ph["date"] = pd.to_datetime(ph["date"])
    coverage = (
        ph.groupby("cusip")
        .agg(n_days=("date", "count"), first=("date", "min"), last=("date", "max"))
        .sort_values("last", ascending=False)
        .reset_index()
    )
    print(f"  {len(ph)} rows, {ph['cusip'].nunique()} CUSIPs\n")
    print(coverage.to_string(index=False))
else:
    print("  No price_history.csv found")

# ── 6. Gap detection ─────────────────────────────────────────────────────────
from bloomberg import _prev_business_day, find_price_gaps
from position_manager import load_all_trades

all_trades = load_all_trades()
print(f"\n=== Gap detection  ({len(all_trades)} total trade rows, {all_trades['cusip'].nunique()} CUSIPs) ===")

end_date = _prev_business_day(date.today())
gaps = find_price_gaps(all_trades)
if gaps:
    print(f"  {len(gaps)} gap ranges across {len({c for c, _, _ in gaps})} CUSIPs:\n")
    for cusip, start, end in gaps:
        print(f"  {cusip}  {start} → {end}  ({(end - start).days} calendar days)")
else:
    print("  No gaps — price_history.csv is complete for all hold intervals.")

# ── 7. Hold intervals per CUSIP ──────────────────────────────────────────────
print(f"\n=== Hold intervals per CUSIP  (end_date={end_date}) ===")
interval_rows: list[dict] = []
for cusip, group in all_trades.groupby("cusip"):
    group = group.sort_values("trade_date").reset_index(drop=True)
    running = 0.0
    istart = None
    for _, row in group.iterrows():
        was_flat = abs(running) < 0.01
        running += float(row["nominal"])
        is_flat = abs(running) < 0.01
        td = row["trade_date"].date() if hasattr(row["trade_date"], "date") else row["trade_date"]
        if was_flat and not is_flat:
            istart = td
        elif not was_flat and is_flat and istart:
            interval_rows.append({"cusip": cusip, "start": istart, "end": td, "held": False})
            istart = None
    if istart:
        interval_rows.append({"cusip": cusip, "start": istart, "end": end_date, "held": True})

if PRICE_HISTORY.exists():
    ph2 = pd.read_csv(PRICE_HISTORY, dtype={"cusip": str})
    ph2["date"] = pd.to_datetime(ph2["date"]).dt.date
    have = set(zip(ph2["cusip"].astype(str), ph2["date"].astype(str)))
else:
    have = set()

summary = []
for r in interval_rows:
    cusip = r["cusip"]
    bdays = [d.date() for d in pd.bdate_range(r["start"], r["end"])]
    total = len(bdays)
    have_n = sum(1 for d in bdays if (cusip, d.isoformat()) in have)
    missing = total - have_n
    status = "✓ complete" if missing == 0 else ("⚠ partial" if have_n > 0 else "✗ missing")
    summary.append({
        "cusip": cusip,
        "start": r["start"],
        "end": r["end"],
        "currently_held": r["held"],
        "bdays_needed": total,
        "have": have_n,
        "missing": missing,
        "status": status,
    })

print(pd.DataFrame(summary).to_string(index=False))

# ── 8. What the import functions actually return ──────────────────────────────
from bloomberg import read_history_from_template, read_prices_from_template, read_static_from_template

print("\n=== read_prices_from_template ===")
try:
    prices = read_prices_from_template(TEMPLATE)
    if prices:
        for cusip, px in prices.items():
            print(f"  {cusip}: {px}")
    else:
        print("  (empty)")
except ValueError as e:
    print(f"  ERROR: {e}")

print("\n=== read_static_from_template ===")
try:
    sdf = read_static_from_template(TEMPLATE)
    print(sdf.to_string() if not sdf.empty else "  (empty)")
except ValueError as e:
    print(f"  ERROR: {e}")

print("\n=== read_history_from_template ===")
try:
    hdf = read_history_from_template(TEMPLATE)
    if not hdf.empty:
        print(f"  {len(hdf)} rows, {hdf['cusip'].nunique()} CUSIPs")
        print(hdf.groupby("cusip").agg(n=("date", "count"), first=("date", "min"), last=("date", "max")))
    else:
        print("  (empty)")
except ValueError as e:
    print(f"  ERROR: {e}")

# ── 9. The file accruals actually reads: data/bonds_static.csv ────────────────
# This is the file that drives bond names + maturity in every dashboard tab.
# If names show as None and accruals warns "missing maturity_date", the blanks
# live HERE — not in the template.
BONDS_STATIC = config.BONDS_STATIC_PATH
print(f"\n=== data/bonds_static.csv  ({BONDS_STATIC}) ===")
if BONDS_STATIC.exists() and BONDS_STATIC.stat().st_size > 0:
    bs = pd.read_csv(BONDS_STATIC, dtype={"cusip": str})
    print(f"  {len(bs)} rows, columns: {list(bs.columns)}\n")
    blank_name = bs["name"].isna() if "name" in bs.columns else pd.Series([], dtype=bool)
    blank_mat = bs["maturity_date"].isna() if "maturity_date" in bs.columns else pd.Series([], dtype=bool)
    print(f"  rows with BLANK name        : {int(blank_name.sum())}")
    print(f"  rows with BLANK maturity_date: {int(blank_mat.sum())}")
    if blank_mat.any():
        print("\n  CUSIPs missing maturity_date (these trigger the accruals warning "
              "AND, in the old code, blocked the entire static save):")
        for c in bs.loc[blank_mat, "cusip"].astype(str):
            print(f"    - {c}")
else:
    print("  NOT FOUND or empty — run generate_sample_data.py / import first.")

# ── 10. Dry-run the merge: would the static save have been blocked? ───────────
print("\n=== Dry-run merge (read_static_from_template → merge) ===")
try:
    from data_io import _validate_bonds, BONDS_COLUMNS

    sdf = read_static_from_template(TEMPLATE)
    if sdf.empty:
        print("  Static sheet read returned no rows — nothing to merge.")
    else:
        tmpl_cusips = set(sdf["cusip"].astype(str))
        print(f"  Template Static resolved {len(tmpl_cusips)} CUSIP(s): {sorted(tmpl_cusips)}")

        if BONDS_STATIC.exists() and BONDS_STATIC.stat().st_size > 0:
            existing = pd.read_csv(BONDS_STATIC, dtype={"cusip": str})
            csv_cusips = set(existing["cusip"].astype(str))
        else:
            existing = pd.DataFrame(columns=BONDS_COLUMNS)
            csv_cusips = set()

        in_csv_not_template = csv_cusips - tmpl_cusips
        if in_csv_not_template:
            print(f"  ⚠ In bonds_static.csv but NOT resolved by template "
                  f"(govt? matured? wrong suffix): {sorted(in_csv_not_template)}")
            print("    → after merge these stay blank → OLD code: save BLOCKED for ALL bonds.")
            print("    → NEW code: kept as placeholders, every other bond saved fine.")
        else:
            print("  ✓ Every bonds_static.csv CUSIP was resolved by the template.")

        # Reproduce the merge and check what strict validation would say.
        existing["cusip"] = existing["cusip"].astype(str)
        fetched = sdf.reindex(columns=BONDS_COLUMNS).copy()
        fetched["cusip"] = fetched["cusip"].astype(str)
        merged = (existing.set_index("cusip")
                  .combine_first(fetched.set_index("cusip")).reset_index())
        if "first_coupon_date" in merged and "maturity_date" in merged:
            m = merged["first_coupon_date"].isna()
            merged.loc[m, "first_coupon_date"] = merged.loc[m, "maturity_date"]
        problems = _validate_bonds(merged)
        if problems:
            print(f"\n  STRICT save would RAISE on {len(problems)} row(s) "
                  f"(this is what silently dropped your static data):")
            for p in problems:
                print(f"    {p}")
        else:
            print("\n  ✓ Strict save would succeed — no blocking rows.")
except ValueError as e:
    print(f"  ERROR (Bloomberg refreshing?): {e}")

# ── 11. Ticker round-trip check ────────────────────────────────────────────
# read_static_from_template/read_prices_from_template recover the CUSIP for
# each row from template_state.json (written by prepare_template). If that
# state is missing/stale they fall back to stripping a suffix off the ticker
# text in col A — which silently derives the WRONG cusip for any bond whose
# bbg_ticker isn't exactly "{cusip} {suffix}" (e.g. a custom govt/Treasury
# description ticker). This section flags any such mismatch directly.
from bloomberg import _cusip_from_ticker, _load_template_state

print("\n=== Ticker round-trip check (template_state.json vs col A) ===")
TEMPLATE_STATE = Path("data/prices/template_state.json")
if TEMPLATE_STATE.exists():
    tstate = json.loads(TEMPLATE_STATE.read_text())
    print(f"  template_state.json found: {TEMPLATE_STATE}")
else:
    tstate = {}
    print("  NOT FOUND — readers will fall back to ticker-text parsing for every row "
          "(run ① Prepare to create it).")

for sheet_name, ws in (("Live MTM", ws_live), ("Static", ws_static)):
    ordered = tstate.get(sheet_name, [])
    print(f"\n  -- {sheet_name} --")
    mismatches = 0
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if raw is None:
            continue
        expected = ordered[r - 2] if (r - 2) < len(ordered) else None
        derived = _cusip_from_ticker(str(raw))
        if expected is not None and expected != derived:
            mismatches += 1
            print(f"    ⚠ row {r}: ticker={raw!r} → derived cusip={derived!r} "
                  f"but template_state says cusip={expected!r} "
                  "(would have been MISFILED under the wrong key without the fix)")
    if mismatches == 0:
        print(f"    no mismatches in {len(ordered) or '0'} mapped row(s)")
